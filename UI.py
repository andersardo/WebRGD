# -*- coding: utf-8 -*-
# This Python file uses the following encoding: utf-8
from collections import OrderedDict
import codecs, time, logging, os.path

from uiUtils import listPersons, listFamilies, personView, familyView, famDisp, persMatchDisp
from uiUtils import listPersonSkillnad, listFamiljeSkillnad, ignoreRelation, getFlags#, addFlag
from utils import setOKfamily, setEjOKfamily, setOKperson, setEjOKperson, split#, kopplaLoss
from uiUtils import dbfind,familyViewAll
from workFlow import workFlowUI, doUpload, cleanUp, getDBselect, listOldLogs
from graphUtils import genGraph
from relationEdit import editList, viewChildErr, viewPartnerErr, viewNoRelErr, viewDubbl, viewQueryHits
from errRelationUtils import mergeFam, mergePers
from queryUtils import doQuery
import conf.config, common
#print conf.config

import bottle
from bottle import response
from beaker.middleware import SessionMiddleware
from cork import Cork
from bson.objectid import ObjectId

# Use users.json and roles.json in the local conf directory
aaa = Cork('conf', email_sender='WebRGD@dis.se', smtp_server=conf.config.mailserver)

# alias the authorization decorator with defaults
authorize = aaa.make_auth_decorator(fail_redirect="/login", role="user")

app = bottle.app()
session_opts = {
    'session.cookie_expires': True,
    'session.encrypt_key': 'please use a random key and keep it secret!',
    'session.httponly': True,
    'session.timeout': 3600 * 24,  # 1 day
    'session.type': 'cookie',
    'session.validate_key': True,
}
app = SessionMiddleware(app, session_opts)

#from bottle import Bottle, run, static_file, request, template
#app = Bottle()
#actionLog = open('actions.log', 'a')

@bottle.route('/static/<filepath:path>')
def server_static(filepath):
#    return bottle.static_file(filepath, root='/home/anders/SVNprojects/staff.it-aar.RGDmongo/static')
    return bottle.static_file(filepath, root='static')

@bottle.route('/img/<filepath:path>')
def server_static(filepath):
#    return bottle.static_file(filepath, root='/home/anders/SVNprojects/staff.it-aar.RGDmongo/static')
    return bottle.static_file(filepath, root='static')

@bottle.hook('before_request')
def setupCommon():
    from datetime import datetime
    bottle.request.session = bottle.request.environ['beaker.session']
    try:
        user = aaa.current_user.username
    except:
        user = 'NoUser'
    workDB = bottle.request.query.workDB
    if (not workDB) and ('workDB' in bottle.request.session):
        workDB = bottle.request.session['workDB']
    matchDB = bottle.request.query.matchDB
    if (not matchDB) and ('matchDB' in bottle.request.session):
        matchDB = bottle.request.session['matchDB']
    if workDB:
        #KOLLA possibly store mongoClient in session? Param to init?
        common.config = common.init(workDB, matchDBName = matchDB)
        bottle.request.session['workDB'] = workDB
        bottle.request.session['matchDB'] = matchDB
        bottle.request.session.save()
        if conf.config.logging and ('action' in bottle.request.url or
                                    'runProg' in bottle.request.url):
            rec = {'type': 'admin', 'time': time.time(), 'workDB': workDB,
                   'matchDB': matchDB, 'url': bottle.request.url,
                   'from': bottle.request.remote_addr, 'user': user}
            common.config['originalData'].insert_one(rec)
    bottle.response.set_header("Cache-Control", "no-cache")
    #print user, bottle.request.remote_addr, str(datetime.now()), bottle.request.url
    logging.info('%s %s %s', user, bottle.request.remote_addr, bottle.request.url)

@bottle.route('/')
@authorize()
def index():
    if 'directory' in bottle.request.session:
        (files, dbs, workingDir, activeUser) = workFlowUI(aaa.current_user.username, bottle.request.session['directory'])
    else:
        (files, dbs, workingDir, activeUser) = workFlowUI(aaa.current_user.username, None)
        bottle.request.session['directory'] = workingDir
    bottle.request.session['activeUser'] = activeUser
    bottle.request.session.save()
    return bottle.template('index', dbs=dbs, files=files, message = None, role=aaa.current_user.role)

#Workflow actions
@bottle.route('/workflow/upload', method='POST')
@authorize()
def upload():
    (mess, fn) = doUpload(bottle.request.session['directory'], bottle.request.files.get('gedcomfile'))
    logging.info(mess)
    (files, dbs, tmp, tmpu) = workFlowUI(aaa.current_user.username, bottle.request.session['directory'])
    return bottle.template('index', dbs=dbs, files=files, message = mess, role=aaa.current_user.role)

@bottle.route('/workflow/combined', method='POST')
@authorize()
def combined():
    (mess, fdir, fn) = doUpload(bottle.request.session['directory'], bottle.request.files.get('gedcomfile'))
    yield '<meta http-equiv="Content-Type" content="text/html; charset=UTF-8" />'
    yield '<meta charset="UTF-8" />'
    yield '<pre>'+mess.encode('UTF-8')+"\n"
    fn = fdir + '/' + fn
    cmd = ['python', 'indataValidering.py', bottle.request.session['directory'], fn]
    if bottle.request.forms.get('resmail'):
        cmd.append('--email')
        cmd.append(aaa.current_user.email_addr)
    if bottle.request.forms.get('namn'): cmd.append('--namn')
    if bottle.request.forms.get('ort'): cmd.append('--ort')
    if bottle.request.forms.get('dubl'): cmd.append('--dubl')
    if bottle.request.forms.get('sour'): cmd.append('--sour')
    import subprocess
    yield "\nAvvakta programmet:  " + cmd[1] + "<br>\n"
    from time import sleep
    import tempfile
    logf = tempfile.TemporaryFile(dir = bottle.request.session['directory'])
    job = subprocess.Popen(cmd,stdout=logf,stderr=subprocess.STDOUT)
    while job.poll() is None:
            #get Log
            yield '.'
            sleep(2)
    yield '</pre>'
    outerr = job.returncode
    logf.seek(0)
#    outdata = logf.read().decode("utf-8")
    outdata = logf.read()

    if outerr:
        yield '<h1>Exekverings **FEL**</h1>' + "\n"
        yield '<pre>' + str(outerr) + '</pre>'
        errf = open( fdir + '/indataValidering.err', 'wb')
        errf.write(str(outerr))
        errf.close()
    yield '<h2>Log</h2>'
    yield '<pre>' + outdata + '</pre>'
    logfile = codecs.open( fdir + '/indataValidering.log', 'wb', 'utf-8')
    logfile.write(outdata.decode("utf-8"))
    logfile.close()

    cmd = ['python', 'importGedcom.py', bottle.request.session['activeUser'],
              fn+'_UTF8']
#    actionLog.write(' '.join(cmd)+"\n")
    yield "\nAvvakta programmet:  " + cmd[1] + "<br>\n"
#    job = subprocess.Popen(cmd,stdout=subprocess.PIPE,stderr=subprocess.PIPE)
    logf = tempfile.TemporaryFile(dir = bottle.request.session['directory'])
    job = subprocess.Popen(cmd,stdout=logf,stderr=subprocess.STDOUT)
    while job.poll() is None:
            #get Log
            yield '.'
            sleep(2)
#    (outdata,outerr) = job.communicate()
    yield '</pre>'
    outerr = job.returncode
    logf.seek(0)
    outdata = logf.read()
    yield '<br><a href="/">Tillbaka till startsida</a>'
    if outerr:
            yield '<h1>Exekverings **FEL**</h1>' + "\n"
            yield '<pre>' + str(outerr) + '</pre>'
            errf = open( fdir + '/importGedcom.err', 'wb')
            errf.write(str(outerr))
            errf.close()
    yield '<h2>Log</h2>'
    yield '<pre>' + outdata + '</pre>'
    yield 'Klar' + "\n"
    yield '</pre>'
    logfile = codecs.open( fdir + '/importGedcom.log', 'wb', 'utf-8')
    logfile.write(outdata)
    logfile.close()

@bottle.route('/workflow/download')
@authorize()
def download():
    import subprocess
    filn = 'RGD_' + bottle.request.query.workDB + '.GED'
    response.add_header('Expires', '0')
    response.add_header('Cache-Control', "must-revalidate, post-check=0, pre-check=0")
    response.set_header('Content-Type', "application/force-download")
    response.add_header('Content-Type', "application/octet-stream")
    response.add_header('Content-Type', "application/download")
    response.add_header('Content-Disposition', 'attachment; filename='+filn)
    response.add_header('Content-Transfer-Encoding', 'binary')
    return subprocess.check_output(['python','exportGedcom.py', bottle.request.query.workDB])

@bottle.route('/getFile')
@authorize()
def getfile():
    fn = bottle.request.query.fil
    if ( ( not fn.startswith(bottle.request.session['directory']) ) or
         ( not os.path.isfile(fn) ) or  ( '..' in fn ) ):
        return 'File not found: ' + os.path.basename(fn)
    if fn.endswith('.zip'):
        response.add_header('Expires', '0')
        response.add_header('Cache-Control', "must-revalidate, post-check=0, pre-check=0")
        response.set_header('Content-Type', "application/force-download")
        response.add_header('Content-Type', "application/octet-stream")
        response.add_header('Content-Type', "application/download")
        response.add_header('Content-Disposition', 'attachment; filename=all.zip')
        response.add_header('Content-Transfer-Encoding', 'binary')
        f = codecs.open(fn, "r")
        mess = f.read()
        f.close()
	return mess
    elif fn.endswith('.CSV'):
        response.add_header('Expires', '0')
        response.add_header('Cache-Control', "must-revalidate, post-check=0, pre-check=0")
        response.set_header('Content-Type', "application/force-download")
        response.add_header('Content-Type', "application/octet-stream")
        response.add_header('Content-Type', "application/download")
        response.add_header('Content-Disposition', 'attachment; filename=RGDK.CSV')
        response.add_header('Content-Transfer-Encoding', 'binary')
        f = codecs.open(fn, "r")
        mess = f.read()
        f.close()
    elif os.path.basename(fn) == 'DgDub.txt':
        response.add_header('Expires', '0')
        response.add_header('Cache-Control', "must-revalidate, post-check=0, pre-check=0")
        response.set_header('Content-Type', "application/force-download")
        response.add_header('Content-Type', "application/download")
        response.add_header('Content-Type', "text/plain")
        response.add_header('Content-Disposition', 'attachment; filename=DgDub.txt')
        f = codecs.open(fn, "r")
        mess = f.read()
        f.close()
	return mess
    else:
        response.set_header('Content-Type', "text/html; charset=UTF-8")
        f = codecs.open(fn, "r", "utf-8")
#        mess = '<pre>' + f.read() + '</pre>'
        mess = f.read()
        f.close()
    return bottle.template('getfile', file = ' '.join(fn.rsplit('/',2)[1:]), message = mess)

@bottle.route('/importBidrag')
@authorize()
def importidrag():
    #dir = bottle.request.query.dir
    return runprog('bidrag')

@bottle.route('/runProg/<prog>')
@authorize()
def runprog(prog):
    import subprocess
    #prog in indatavalidation, import, match, merge
    cmd = None
    mess = ''
#FIX verify bottle.request.query.file, workDB, matchDB
    if prog == 'import':
        cmd = ['python', 'importGedcom.py', bottle.request.session['activeUser'],
               bottle.request.session['directory']+'/'+bottle.request.query.file]
    elif prog == 'bidrag':
       #Function to import RGD real contributions
       # strip leading path from file name to avoid directory traversal attacks
       fndir = os.path.basename(bottle.request.query.dir)
       fdir = bottle.request.session['directory'] + '/' + fndir
       if os.path.isdir(fdir):
           import shutil
           shutil.rmtree(fdir)
       try:
           os.mkdir(fdir)
       except Exception, e:
           message += str(e)
       bottle.request.query.workDB = bottle.request.session['activeUser']+'_'+fndir
       cmd = ['python', 'importBidrag.py', bottle.request.session['activeUser'],
               bottle.request.session['directory'], bottle.request.query.dir ]
    elif prog == 'match':
        #check parameters FIX
        if bottle.request.query.workDB == bottle.request.query.matchDB:
#            cmd = ['python', 'matchDubl.py', bottle.request.query.workDB, bottle.request.query.matchDB]
            cmd = ['python', 'MatchDubl.py', bottle.request.query.workDB, bottle.request.query.matchDB, bottle.request.session['directory']]
        else:
            cmd = ['python', 'match.py', bottle.request.query.workDB, bottle.request.query.matchDB]
#            if bottle.request.session['activeUser'] == 'tester':
#                cmd = ['python', 'matchFam.py', bottle.request.query.workDB, bottle.request.query.matchDB]
    elif prog == 'matchDublExp':
        cmd = ['python', 'MatchDubl.py', bottle.request.query.workDB, bottle.request.query.matchDB, bottle.request.session['directory']]
    elif prog == 'xldubl':
        cmd = ['python', 'runDubbtestx.py', bottle.request.query.workDB, bottle.request.session['directory']]
    elif prog == 'sanity':
        #check parameters FIX
        cmd = ['python', 'sanityCheck.py', bottle.request.query.workDB, bottle.request.query.matchDB]
    elif prog == 'indatavalidation':
        cmd = ['python', 'indataValidering.py', bottle.request.session['directory'],
               bottle.request.session['directory']+'/'+bottle.request.query.file]
        if 'resmail' in bottle.request.query:
               cmd.append('--email')
               cmd.append(bottle.request.session['activeUser'])
        if 'namn' in bottle.request.query: cmd.append('--namn')
        if 'ort' in bottle.request.query: cmd.append('--ort')
        if 'dubl' in bottle.request.query: cmd.append('--dubl')
    elif prog == 'merge':
        #check parameters FIX
        cmd = ['python', 'merge.py', bottle.request.query.workDB, bottle.request.query.matchDB]
    else:
        mess = prog + ' Not implemented'
    if cmd:
        (user,db) = bottle.request.query.workDB.split('_', 1)
        fdir = './files/'+user+'/'+db+'/'
        mess = 'Running ' + ' '.join(cmd)
        mess = 'Running ' + ' ' + cmd[1]
        yield "Avvakta programmet:  " + cmd[1] + "<br>\n"
        from time import sleep
        import tempfile
        logf = tempfile.TemporaryFile(dir = bottle.request.session['directory'])
        job = subprocess.Popen(cmd,stdout=logf,stderr=subprocess.STDOUT)
        while job.poll() is None:
            #get Log?
            yield '.'
            sleep(2)
        outerr = job.returncode
        logf.seek(0)
        outdata = logf.read()
        logf.close()
        yield '<br><a href="/">Tillbaks till startsida</a>'
        if outerr:
            yield '</pre><h1>Exekverings **FEL**</h1>' + "\n"
            yield '<pre>' + str(outerr) + '</pre>'
            mess += '<h1>Exekverings **FEL**</h1>'
            mess += '<pre>' + str(outerr) + '</pre>'
            errf = open( fdir + prog + '.err', 'wb')
            errf.write(str(outerr))
            errf.close()
        yield '<h2>Log</h2>'
        yield '<pre>' + outdata + '</pre>'
        logfile = codecs.open( fdir + prog + '.log', 'wb', 'utf-8')
        logfile.write(outdata.decode("utf-8"))
        logfile.close()
    else:
        (files, dbs, tmp, tmpu) = workFlowUI(aaa.current_user.username, bottle.request.session['directory'])
        yield bottle.template('index', dbs=dbs, files=files, message = mess, role=aaa.current_user.role)

#Manual matchning
@bottle.route('/list/<typ>')
@authorize()
#typ in 'persons', 'families'
def lists(typ):
    filter = OrderedDict()
    for stat in (common.statOK, common.statManuell, common.statEjOK, ['FamEjOK']):
        for s in list(stat): filter[s] = ''
    multi = {'DB1': '', 'DB2': ''}
    for m in bottle.request.params.getall('multi'):
        multi[m]='checked'
        for s in list(common.statOK): filter[s] = 'checked'
        for s in list(common.statManuell): filter[s] = 'checked'
    filtergroup = OrderedDict()
    filtergroup['allOK'] = ''
    filtergroup['allManuell'] = ''
    filtergroup['allEjOK'] = ''
    page = int(bottle.request.params.pageNo or '1')
    prevnext = bottle.request.params.page or ''
    if prevnext == 'prev': page += -1
    elif prevnext == 'next': page += 1
    else: page = 1
    if page <= 0: page = 1
    tot = ''
    for f in bottle.request.params.getall('filters'):
        filter[f]='checked'
    for f in bottle.request.params.getall('filtergroups'):
        filtergroup[f] = 'checked'
        if f == 'allOK':
            for s in list(common.statOK): filter[s] = 'checked'
        elif f == 'allManuell':
            for s in list(common.statManuell): filter[s] = 'checked'
        elif f == 'allEjOK':
            for s in list(common.statEjOK)+['FamEjOK']: filter[s] = 'checked'
    dbs = bottle.request.session['workDB'] + ' -> ' + bottle.request.session['matchDB']
    if typ == 'persons':
        tit = 'Personlista: ' + dbs
        (rows, tot) = listPersons(filter, multi, page = page)
    elif typ == 'families':
        tit = 'Familjelista: ' + dbs
        (rows, tot) = listFamilies(filter, multi, page = page)
    else:
        tit = '???'
        rows = []
    return bottle.template('lists', typ = typ, title = tit, page = page, tot = tot,
                    filtergroup=filtergroup, filter=filter, multi=multi, prow=rows)

@bottle.route('/search')
@authorize()
def search():
    gedid = bottle.request.query.gedid
    if bottle.request.query.databas == 'work':
        p = common.config['persons'].find_one({'refId': gedid}, {'_id': 1})
        if p:
            res = lists('persons')
            bottle.request.query.wid = p['_id']
            return res + views('persons')
    else:
        p = common.config['match_persons'].find_one({'refId': gedid}, {'_id': 1})
        if p:
            res = lists('persons')
            bottle.request.query.mid = p['_id']
            return res + views('persons')
    if bottle.request.query.databas == 'work':
        f = common.config['families'].find_one({'refId': gedid}, {'_id': 1})
        if f:
            res = lists('families')
            bottle.request.query.wid = f['_id']
            return res + views('families')
    else:
        f = common.config['match_families'].find_one({'refId': gedid}, {'_id': 1})
        if f:
            res = lists('families')
            bottle.request.query.mid = f['_id']
            return res + views('families')
    return lists('families')

#list suspected dublicates
@bottle.route('/listDubl')
@authorize()
def listdubl():
    if not bottle.request.query.workDB:
        if 'workDB' in bottle.request.session:
            bottle.request.query.workDB = bottle.request.session['workDB']
        else:
            return 'Databas I ej vald - programmet avslutas<br><a href="/">Tillbaka till startsida</a>'
    bottle.request.query.matchDB = bottle.request.query.workDB
    bottle.request.session['matchDB'] = bottle.request.query.workDB
    #need to re-init database collections
    common.config = common.init(bottle.request.query.workDB,
                                matchDBName = bottle.request.query.matchDB)
    #test if workDB matched against itself
    dbOK = getDBselect('listDubl', bottle.request.query.workDB,
             bottle.request.session['activeUser'], bottle.request.session['directory'])
    if dbOK == 'No valid choices':
        #if not run match
        return runprog('match')
    #else show list
    from uiUtils import persDisp
    tit = 'Lista från alternativ dubblettkontroll'
    if bottle.request.params.sortNS:
        sortVal = 'checked'
        sorting = [('nodesim', -1),('sortDubl', -1)]
    else:
        sortVal = ''
        sorting = [('sortDubl', -1),('nodesim', -1)]
    page = int(bottle.request.params.pageNo or '1')
    prevnext = bottle.request.params.page or ''
    if prevnext == 'prev': page += -1
    elif prevnext == 'next': page += 1
    else: page = 1
    if page <= 0: page = 1
    rows = [['#',u'Namn/refId', u'Född', u'Död','Score/NodeSim', u'Namn/refId', u'Född', u'Död','Visa']]
    i = (page-1)*10
    args = {'where': 'visa', 'what': '/view/persons', 'buttons': 'No'}
#    for mt in common.config['matches'].find({'nodesim': {'$gt': 0.3}}).sort([('nodesim', -1),('sortDubl', -1)]).limit( 50 ):
    tot = common.config['matches'].find({'nodesim': {'$gt': 0.3}}).count()
    for mt in common.config['matches'].find({'nodesim': {'$gt': 0.3}}).sort(sorting).skip((page-1)*10).limit( 10 ):
        #print mt['pwork']['name'],mt['pmatch']['name'], mt['sortDubl'], mt['nodesim']
        i += 1
        row = [str(i)]
        row.extend(persDisp(mt['pwork']))
        row.append(str(mt['sortDubl'])+'<br>'+str(mt['nodesim']))
        row.extend(persDisp(mt['pmatch']))
        args['wid'] = str(mt['workid'])
        args['mid'] = str(mt['matchid'])
        row.append('<button onclick="doAction('+str(args)+')">Visa</button>')
        rows.append(row)
    return bottle.template('dubl', title = tit, page = page, tot = tot, 
                           prow=rows, sort=sortVal)

#list suspected dublicates
@bottle.route('/listDublExp')
@authorize()
def listdublexp():
    if not bottle.request.query.workDB:
        bottle.request.query.workDB = bottle.request.session['workDB']
    bottle.request.query.matchDB = bottle.request.query.workDB
    bottle.request.session['matchDB'] = bottle.request.query.workDB
    #need to re-init database collections
    common.config = common.init(bottle.request.query.workDB,
                                matchDBName = bottle.request.query.matchDB)
    #test if workDB matched against itself
    dbOK = getDBselect('listDubl', bottle.request.query.workDB,
             bottle.request.session['activeUser'], bottle.request.session['directory'])
    if dbOK == 'No valid choices':
        #if not run match
        return runprog('matchDublExp')
    #else show list
#Bort
#    import json
#    f = open('DublStat.json', 'rw')
#    dublStat = json.loads(f.read())
#    f.close()

    from uiUtils import persDisp
    tit = 'Lista från alternativ dubblettkontroll'
#    sorting =  [('alg1', -1)]
#    sortalt = {'sortDubl': '', 'kscore': '', 'alg1': '', 'alg2': '', 'alg3': ''}
    sorting =  [('Match', -1)]
    sortalt = {'Match': '', 'XL': '', 'Snitt': ''}
    for m in bottle.request.params.getall('sorting'):
        sortalt[m]='checked'
        sorting = [(m, -1)]
    if sorting == [('Match', -1)]: sortalt['Match']='checked'
    page = int(bottle.request.params.pageNo or '1')
    prevnext = bottle.request.params.page or ''
    if prevnext == 'prev': page += -1
    elif prevnext == 'next': page += 1
    else: page = 1
    if page <= 0: page = 1
#    rows = [['#','stat',u'Namn/refId', u'Född', u'Död','Scores', u'Namn/refId', u'Född', u'Död','Visa']]
    rows = [['#',u'Namn/refId', u'Född', u'Död','Scores', u'Namn/refId', u'Född', u'Död','Visa']]
    i = (page-1)*10
    args = {'where': 'visa', 'what': '/view/persons', 'buttons': 'No'}
#    for mt in common.config['matches'].find({'nodesim': {'$gt': 0.3}}).sort([('nodesim', -1),('sortDubl', -1)]).limit( 50 ):
    tot = common.config['matches'].find({'nodesim': {'$gt': 0.3}}).count()
    for mt in common.config['matches'].find({'nodesim': {'$gt': 0.3}}).sort(sorting).skip((page-1)*10).limit( 10 ):
        #print mt['pwork']['name'],mt['pmatch']['name'], mt['sortDubl'], mt['nodesim']
        i += 1
        key = mt['pwork']['refId']+';'+mt['pmatch']['refId']
#bort
        #print key
#        if key in dublStat: status = dublStat[key]
#        else: status = 'xx'
#        row = [str(i),status]
        row = [str(i)]
        row.extend(persDisp(mt['pwork']))
        r = ''
        for m in sortalt.keys():
            if mt[m] > 0.001:
                r += m + '=' + str(round(mt[m],3)) + '<br>'
#        row.append(str(mt['sortDubl'])+'<br>'+str(mt['nodesim']))
        row.append(r)
        row.extend(persDisp(mt['pmatch']))
        args['wid'] = str(mt['workid'])
        args['mid'] = str(mt['matchid'])
        row.append('<button onclick="doAction('+str(args)+')">Visa</button>')
        rows.append(row)
    return bottle.template('dublExp', title = tit, page = page, tot = tot, 
                           prow=rows, sort=sortalt)

#Skillnader
@bottle.route('/listSkillnad/<typ>')
@authorize()
#typ in 'persons', 'families'
def listSkillnad(typ):
    if bottle.request.query.buttons == 'No': buttons = False
    else: buttons = True
    difftyp = {'Namn': '', 'NamnGrupp': '', 'NamnStavning': '',
               'Datum': '', 'Plats': '', 'Alla': ''}
    diffTypActive = None
    for m in bottle.request.params.getall('difftyp'):
        difftyp[m] = 'checked'
        diffTypActive = m
    page = int(bottle.request.params.pageNo or '1')
    prevnext = bottle.request.params.page or ''
    if prevnext == 'prev': page += -1
    elif prevnext == 'next': page += 1
    else: page = 1
    if page <= 0: page = 1
    if typ == 'persons':
        tit = 'Skillnad Personlista'
        (rows, tot) = listPersonSkillnad(page = page, diffType=diffTypActive)
    elif typ == 'families':
        tit = 'Skillnad Familjelista - Inte implementerat'
        rows = []
        tot = 0
#        (rows,tot) = listFamiljeSkillnad(page = page)
    else:
        tit = '???'
        rows = []
        tot = 0
    return bottle.template('listSkillnad', typ = typ, title = tit, page = page,
                           tot = tot, prow=rows, buttons=buttons, difftyp=difftyp)

#Relation editor
@bottle.route('/relationsEditor/<typ>')
@authorize()
def relationEditor(typ):
    (tit, childErrs, famErrs, relErrs, dubbls) = editList(common.config, typ)
    return bottle.template('listEdit', title = tit, childErrs=childErrs,
                           famErrs=famErrs, relErrs=relErrs, dubbletter=dubbls)

def viewRelErr(person, family, typ):
    if typ == 'child':
        (res, graph) = viewChildErr(person, family.split(':'), common.config)
    elif typ == 'partner':
        (res, graph) = viewPartnerErr(person.split(':'), family, common.config)
    elif typ == 'noRel':
        (res, graph) = viewNoRelErr(person, None, common.config)
    elif typ == 'dubblett':
        (res, graph) = viewDubbl(person, family, common.config)
    elif typ == 'dubblettFind':
        (res, graph) = viewDubbl(person, family, common.config, find=True)
    elif typ == 'queryhits':
        (res, graph) = viewQueryHits(person, common.config)
    else:
        res = [[u'Okänd feltyp', typ, '', '', '']]
        graph = ''
    return bottle.template('viewEdit', rows=res, graph=graph, buttons=None)

@bottle.route('/queryDB')
@authorize()
def queryDB():
    (tit, res) = doQuery(bottle.request.query.q, common.config)
    return bottle.template('listEdit', title = tit, childErrs=[],
                           famErrs=[], relErrs=[], dubbletter=res) #TMP

#Likheter
@bottle.route('/downloadFamMatches')
@authorize()
def downloadFamMatches():
    import StringIO
    output = StringIO.StringIO()
    done = set()
    fileFormat = bottle.request.query.fileFormat
    titleRow = [bottle.request.query.workDB,'','','',bottle.request.query.matchDB,'','']
    if fileFormat == 'xlsx':
        from openpyxl import Workbook
        from openpyxl.writer.write_only import WriteOnlyCell
        from openpyxl.styles import colors
        from openpyxl.styles import Border, Side, Font, Color, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        wb = Workbook(write_only=True)
        ws = wb.create_sheet()
        ws.title = "RGD matches"
        greyFill = PatternFill(start_color='DDDDDD',
                   end_color='DDDDDD',
                   fill_type='solid')
        greenFill = PatternFill(start_color='50FF50',
                   end_color='50FF50',
                   fill_type='solid')
        yellowFill = PatternFill(start_color='FFFF50',
                   end_color='FFFF50',
                   fill_type='solid')
        redFill = PatternFill(start_color='FF5050',
                   end_color='FF5050',
                   fill_type='solid')
        thick = Side(border_style="thick", color="000000")
        thin = Side(border_style="thin", color="000000")
        rowVals = []
        for val in titleRow:
            cell = WriteOnlyCell(ws, value=val.encode("utf-8"))
            cell.font = Font(bold=True)
            rowVals.append(cell)
        ws.append(rowVals)
    else:
        import csv
        CSV =  csv.writer(output, dialect='excel')
        CSV.writerow([s.encode("utf-8") for s in titleRow])
    for fmatch in common.config['fam_matches'].find({'status':
                          {'$in': list(common.statOK.union(common.statManuell))}}):
        rows = famDisp(None, None, fmatch)
        line = False
        lines=0
        for r in rows:
            lines += 1
            if lines == 1: continue
            try:
                k1 = r[1].split('<br/>')[1]
            except:
                k1 = ''
            try:
                k2 = r[5].split('<br/>')[1]
            except:
                k2 = ''
            key = k1+';'+k2
            done.add(key)
            if 'Ignorerad' in r[0]: ignRelI = True
            else: ignRelI = False
            if 'Ignorerad' in r[8]: ignRelII = True
            else: ignRelII = False
            ignFam = False
            #remove html-code for buttons
            try:
                if 'Ignorerad' in r[4]: ignFam = True
                r[4] = r[4].split('<')[0]
                if ign: r[4] += ' Ignorerad'
            except:
                pass
            if r == ['', '', '', '', '', '', '', '', '']:
                line = True
                if fileFormat == 'xlsx': continue
            if fileFormat == 'xlsx':
                rowVals = []
                i=0
                green = False
                yellow = False
                red = False
                #if r[4].endswith(('EjMatch', 'EjOK', 'rEjOK')): red = True
                #elif r[4].endswith(('Manuell', 'rManuell')): yellow = True
                #elif r[4].endswith(('Match', 'OK', 'rOK')): green = True
                if ('EjMatch' in r[4]) or ('EjOK' in r[4]): red = True
                elif ('Match' in r[4]) or ('OK' in r[4]): green = True
                elif 'Manuell' in r[4]: yellow = True
                for val in r[1:8]:
                    i+=1
                    if i == 4: #separator between workDB and matchDB
                        cell = WriteOnlyCell(ws, value=val)
                        cell.border = Border(top=thin, left=thick, right=thick, bottom=thin)
                    else:
                        cell = WriteOnlyCell(ws,
                                value=val.replace('<br/>', "\n").rstrip().encode("utf-8"))
                        cell.alignment = Alignment(wrapText=True)
                        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    if lines <=2: #head
                        cell.font = Font(bold=True)
                        cell.border = Border(top=thick, left=thin, right=thin, bottom=thick)
                        cell.fill = greyFill
                    elif line:
                        if i == 4: #separator between workDB and matchDB
                            cell.border = Border(top=thick, left=thick, right=thick, bottom=thin)
                        else:
                            cell.border = Border(top=thick, left=thin, right=thin, bottom=thin)
                    if (i == 4) and ignFam: cell.fill = redFill
                    elif (i==1) and ignRelI: cell.fill = redFill
                    elif (i==8) and ignRelII: cell.fill = redFill
                    elif green: cell.fill = greenFill
                    elif yellow:  cell.fill = yellowFill
                    elif red:  cell.fill = redFill
                    rowVals.append(cell)
                #line = False
                ws.append(rowVals)
            else:
                CSV.writerow([s.replace('<br/>', "\n").rstrip().encode("utf-8") for s in r])
        if fileFormat != 'xlsx':
            CSV.writerow(['#####', '#####', '#####', '#####', '#####', '|', '#####', '#####', '#####'])
    #List matched persons without families
    for persmatch in common.config['matches'].find({'status':
                          {'$in': list(common.statOK.union(common.statManuell))}}):
        if persmatch['pwork']['refId']+';'+persmatch['pmatch']['refId'] in done: continue
        rows = []
        rows.append(['', u'Namn/refId', u'Född', u'Död', '', u'Namn/refId', u'Född', u'Död', ''])
        rows.append(persMatchDisp('Person', persmatch))
        head = True
        for r in rows:
            #remove html-code for buttons
            try:
                r[4] = r[4].split('<')[0]
            except:
                pass
            if fileFormat == 'xlsx':
                rowVals = []
                i=0
                green = False
                yellow = False
                red = False
                if r[4].endswith(('EjMatch', 'EjOK', 'rEjOK')): red = True
                elif r[4].endswith(('Manuell', 'rManuell')): yellow = True
                elif r[4].endswith(('Match', 'OK', 'rOK')): green = True
                for val in r[1:8]:
                    i+=1
                    if i==4: #separator between workDB and matchDB
                        cell = WriteOnlyCell(ws, value=val)
                        cell.border = Border(top=thin, left=thick, right=thick, bottom=thin)
                    else:
                        cell = WriteOnlyCell(ws,
                                value=val.replace('<br/>', "\n").rstrip().encode("utf-8"))
                        cell.alignment = Alignment(wrapText=True)
                        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    if head:
                        cell.font = Font(bold=True)
                        cell.border = Border(top=thick, left=thin, right=thin, bottom=thick)
                        cell.fill = greyFill
                    if green: cell.fill = greenFill
                    elif yellow:  cell.fill = yellowFill
                    elif red:  cell.fill = redFill
                    rowVals.append(cell)
                head = False
                line = False
                ws.append(rowVals)
            else:
                CSV.writerow([s.replace('<br/>', "\n").rstrip().encode("utf-8") for s in r])
        if fileFormat != 'xlsx':
            CSV.writerow(['#####', '#####', '#####', '#####', '#####', '|', '#####', '#####', '#####'])
    #
    if fileFormat == 'xlsx':
        #doesn't really work?
        for i, column_width in enumerate([8,10,25,25,25,1,25,25,25]):
            ws.column_dimensions[get_column_letter(i+1)].width = (column_width + 2) * 1.2
        wb.save(output)
    #Download
    filn = 'RGD_' + bottle.request.query.workDB + '-' + bottle.request.query.matchDB + '.' + fileFormat
    response.add_header('Expires', '0')
    response.add_header('Cache-Control', "must-revalidate, post-check=0, pre-check=0")
    response.set_header('Content-Type', "application/force-download")
    response.add_header('Content-Type', "application/octet-stream")
    response.add_header('Content-Type', "application/download")
    response.add_header('Content-Disposition', 'attachment; filename='+filn)
    response.add_header('Content-Transfer-Encoding', 'binary')
    return output.getvalue()

@bottle.route('/viewNext')
@bottle.route('/viewNext/<skip>')
@authorize()
def viewNext(skip = None):
    if skip:
        bottle.request.session['skipNo'] += 1 
        skipNo = bottle.request.session['skipNo']
    else:
        bottle.request.session['skipNo'] = 0
        skipNo = 0
    bottle.request.session.save()
    res = ''
    buttons = True
    resL = familyViewAll(skipNo)
    for (prow, wpid, mpid, tab, wfid, mfid) in resL:
        res += bottle.template('person', prow=prow, wid=wpid, mid=mpid, buttons=buttons)
        res += '<h2>Familj</h2>'
        res += bottle.template('family', rows=tab, wfid=wfid, mfid=mfid, buttons=buttons)
    return lists('families') + res

@bottle.route('/view/<typ>')
@authorize()
def views(typ):
    res = ''
    if bottle.request.query.buttons == 'No': buttons = False
    else: buttons = True
    if typ == 'persons':
        resL = personView(bottle.request.query.wid, bottle.request.query.mid)
    elif typ == 'families':
        resL = familyView(bottle.request.query.wid, bottle.request.query.mid)  #How do a list? KOLLA
    elif typ == 'relErr':
        return viewRelErr(bottle.request.query.person, bottle.request.query.family,
                          bottle.request.query.typ)
    elif typ == 'flags':
        flags = getFlags(bottle.request.query.wid, bottle.request.query.fid,
                         role = bottle.request.query.role)
        return bottle.template('flags', flagList=flags,
                               personId=bottle.request.query.wid,
                               famId=bottle.request.query.fid)
    else:
        #resL = []
        return typ + ' - Ej implementerad'
    for (prow, wpid, mpid, tab, wfid, mfid) in resL:
        res += bottle.template('person', prow=prow, wid=wpid, mid=mpid, buttons=buttons)
        res += '<h2>Familj</h2>'
#        print 'P', wpid, mpid, 'F', wfid, mfid
        res += bottle.template('family', rows=tab, wfid=wfid, mfid=mfid, buttons=buttons)
    return res

@bottle.route('/graph')
@authorize()
def showGraph():
    res = genGraph(bottle.request.query.wid, 
                    common.config['families'], common.config['persons'], common.config['relations'],
                    bottle.request.session['directory'], 'Work family')
    try:
        res += genGraph(bottle.request.query.mid, 
                    common.config['match_families'], common.config['match_persons'], common.config['match_relations'],
                    bottle.request.session['directory'], 'Match family')
    except: pass
    return res

@bottle.route('/actions/setOK/family')
@authorize()
def action():
    return setOKfamily(bottle.request.query.wid, bottle.request.query.mid)

@bottle.route('/actions/setOK/person')
@authorize()
def act1():
    return setOKperson(ObjectId(bottle.request.query.wid),
                       ObjectId(bottle.request.query.mid), button=True)

@bottle.route('/actions/setEjOK/family')
@authorize()
def act2():
    return setEjOKfamily(bottle.request.query.wid, bottle.request.query.mid)

@bottle.route('/actions/setEjOK/person')
@authorize()
def act3():
    return setEjOKperson(bottle.request.query.wid, bottle.request.query.mid)

"""
@bottle.route('/actions/kopplaLoss')
@authorize()
def act4():
    return kopplaLoss(bottle.request.query.wid, bottle.request.query.fid, bottle.request.query.role)
"""
@bottle.route('/actions/split')
@authorize()
def actSplit():
    return split(bottle.request.query.wid, bottle.request.query.mid)

@bottle.route('/actions/addFlag')
@authorize()
def act5():
    return addFlag(bottle.request.query.wid, bottle.request.query.fid, bottle.request.query.fltext)

@bottle.route('/actions/ignoreRelation')
@authorize()
def act5():
    return ignoreRelation(bottle.request.query.pid, bottle.request.query.fid,
                          bottle.request.query.role,
                          bottle.request.query.workFam, bottle.request.query.matchFam)

@bottle.route('/actions/mergeFam')
@authorize()
def act5():
    print 'mergeFam', bottle.request.query.id1, bottle.request.query.id2
    return mergeFam(bottle.request.query.id1, bottle.request.query.id2,
                    common.config['persons'], common.config['families'],
                    common.config['relations'], common.config['originalData'])

@bottle.route('/actions/mergePers')
@authorize()
def act5():
    print 'mergePers', bottle.request.query.id1, bottle.request.query.id2
    return mergePers(bottle.request.query.id1, bottle.request.query.id2,
                    common.config['persons'], common.config['families'],
                    common.config['relations'], common.config['originalData'])

@bottle.route('/actions/delRelation')
@authorize()
def act5():
    print 'delRelation', bottle.request.query.id1, bottle.request.query.id2
    common.config['relations'].delete_one({'persId': bottle.request.query.id1,
                                    'famId': bottle.request.query.id2})
    return

#################ADMIN################
@bottle.route('/oldLogs')
@authorize()
def oldlogs():
    mess = listOldLogs(bottle.request.session['activeUser'], bottle.request.query.workDB)
    (tmpf, dbs, tmpwd, tmpu) = workFlowUI(aaa.current_user.username, bottle.request.session['directory'])
    return bottle.template('dbadmin', dbs=dbs, message = mess,
                           user=bottle.request.session['activeUser'],
                           role=aaa.current_user.role)

@bottle.route('/DBbrowse')
@authorize()
def dbbrowse():
    return bottle.template('dbbrowse', db=bottle.request.query.workDB, message=None,
            collections=common.admClient[bottle.request.query.workDB].collection_names())

@bottle.route('/DBadmin')
@authorize()
def dbadmin():
    (tmpf, dbs, tmpwd, tmpu) = workFlowUI(aaa.current_user.username, bottle.request.session['directory'])
    return bottle.template('dbadmin', dbs=dbs, message = None,
                           user=bottle.request.session['activeUser'],
                           role=aaa.current_user.role)

@bottle.route('/DBaction')
@authorize()
def dbaction():
    #Validate paramets FIX
    if not bottle.request.query.workDB:
        mess = 'Ingen databas vald'
    elif bottle.request.query.action == 'del':
        mess = common.deleteDB(bottle.request.query.workDB)
    elif bottle.request.query.action == 'rmMatch':
        mess = common.rmMatchData(bottle.request.query.workDB)
    elif bottle.request.query.action == 'info':
        mess = common.infoDB(bottle.request.query.workDB)
    elif bottle.request.query.action in ('findall', 'findone'):
        mess = dbfind(bottle.request.session['workDB'], bottle.request.query)
        return bottle.template('dbbrowse', db=bottle.request.session['workDB'], message=mess,
            collections=common.admClient[bottle.request.session['workDB']].collection_names())
    else:
        mess = 'Inget att göra ...'
    (tmpf, dbs, tmpwd, tmpu) = workFlowUI(aaa.current_user.username, bottle.request.session['directory'])
    return bottle.template('dbadmin', dbs=dbs, message = mess,
                           user=bottle.request.session['activeUser'],
                           role=aaa.current_user.role)

#########################DEBUG
@bottle.route('/DBdebug')
@authorize()
def dbdebug():
    import debugUtils
    if not bottle.request.query.workDB or not bottle.request.query.matchDB:
        mess = 'Ingen databas vald'
    #workFam = bottle.request.query.workFam
    #matchFam = bottle.request.query.matchFam
    workFam = common.config['families'].find_one({'refId': bottle.request.query.workFam})
    matchFam = common.config['match_families'].find_one({'refId': bottle.request.query.matchFam})
    match = debugUtils.DmatchFam(workFam['_id'], matchFam['_id'], common.config)
    tab = famDisp(workFam['_id'], matchFam['_id'], match)
    return bottle.template('family', rows=tab, wfid=workFam['_id'], mfid=matchFam['_id'], buttons=None)

######################TEST
@bottle.route('/databases/<what>/<DB1>')
def dbs(what, DB1):
    #print bottle.request.session
    #return 'GOT', what, '.', DB
    return getDBselect(what, DB1, bottle.request.session['activeUser'], bottle.request.session['directory'])

    sel="""<select name="matchDB">
<option value="">Select database II</option>
    <option>DB1</option>'
    <option>DBA</option>'
    <option>DBB</option>'
</select>
"""
    return sel
#TEST

###################AUTH###################

def postd():
    return bottle.request.forms

def post_get(name, default=''):
    return bottle.request.POST.get(name, default).strip()

@bottle.post('/login')
def login():
    """Authenticate users"""
    username = post_get('username')
    password = post_get('password')
    if username == 'guest':
        password = ''
    aaa.login(username, password, success_redirect='/', fail_redirect='/login')

@bottle.route('/logout')
def logout():
    #cleanup - if guest: directory - databases
    try:
        if bottle.request.session['username'] == 'guest':
            cleanUp(bottle.request.session['activeUser'], bottle.request.session['directory'])
    except: pass
    aaa.logout(success_redirect='/login')

@bottle.post('/reset_password')
def send_password_reset_email():
    """Send out password reset email"""
    if post_get('username') == bottle.request.session['activeUser']:
        aaa.send_password_reset_email(
            username=post_get('username'),
            email_addr=post_get('email_address',),
            subject='RGD Web-services nytt password'
            )
        return u'Email med reset-kod har skickats. OBS du måste vara inloggad när den används!'
    else:
        sorry_page()

@bottle.route('/change_password/:reset_code')
@authorize(role="user", fail_redirect='/sorry_page')
@bottle.view('password_change_form')
def change_password(reset_code):
    """Show password change form"""
    return dict(reset_code=reset_code)

@bottle.post('/change_password')
def change_password():
    """Change password"""
    aaa.reset_password(post_get('reset_code'), post_get('password'))
    return 'Thanks. <a href="/login">Go to login</a>'

@bottle.route('/admin')
@authorize(role="admin", fixed_role=True, fail_redirect='/sorry_page')
@bottle.view('admin_page')
def admin():
    """Only admin users can see this"""
    #aaa.require(role='admin', fail_redirect='/sorry_page')
    return dict(
        current_user = aaa.current_user,
        users = aaa.list_users(),
        roles = aaa.list_roles()
    )

@bottle.post('/create_user')
@authorize(role="admin", fixed_role=True, fail_redirect='/sorry_page')
def create_user():
    try:
        aaa.create_user(postd().username, postd().role, postd().password, email_addr=postd().email)
        return dict(ok=True, msg='')
    except Exception, e:
        return dict(ok=False, msg=e.message)

@bottle.post('/delete_user')
@authorize(role="admin", fixed_role=True, fail_redirect='/sorry_page')
def delete_user():
    try:
        aaa.delete_user(post_get('username'))
        return dict(ok=True, msg='')
    except Exception, e:
        logging.ERROR(repr(e))
        return dict(ok=False, msg=e.message)

@bottle.post('/create_role')
@authorize(role="admin", fixed_role=True, fail_redirect='/sorry_page')
def create_role():
    try:
        aaa.create_role(post_get('role'), post_get('level'))
        return dict(ok=True, msg='')
    except Exception, e:
        return dict(ok=False, msg=e.message)

@bottle.post('/delete_role')
@authorize(role="admin", fixed_role=True, fail_redirect='/sorry_page')
def delete_role():
    try:
        aaa.delete_role(post_get('role'))
        return dict(ok=True, msg='')
    except Exception, e:
        return dict(ok=False, msg=e.message)

# Static pages

@bottle.route('/login')
@bottle.view('login_form')
def login_form():
    """Serve login form"""
    return {}

@bottle.route('/sorry_page')
def sorry_page():
    """Serve sorry page"""
    return '<p>Sorry, you are not authorized to perform this action</p>'

########################################
logging.basicConfig(level=logging.INFO,
    format = '%(levelname)s %(asctime)s %(module)s:%(funcName)s:%(lineno)s - %(message)s')

try:
    conf.config.host
except AttributeError:
    logging.info('Getting local IP-no')
    import socket
    conf.config.host = socket.gethostbyname(socket.gethostname())
try:
    conf.config.port
except:
    logging.info('Using default port 8085')
    conf.config.port = 8085

bottle.run(app, debug=True, host=conf.config.host, server=conf.config.wsgiserver,
           port=conf.config.port, reloader=True)

